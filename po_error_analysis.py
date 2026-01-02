import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime
import openpyxl
import xml.etree.ElementTree as ET
from xml.dom import minidom
import copy

# Configure Streamlit
st.set_page_config(page_title="PO Migration Error Analysis", layout="wide")

# Store original XML for later modification
if 'original_xml_content' not in st.session_state:
    st.session_state.original_xml_content = None
if 'modified_dataframes' not in st.session_state:
    st.session_state.modified_dataframes = {}
if 'header_mappings' not in st.session_state:
    st.session_state.header_mappings = {}
if 'edit_history' not in st.session_state:
    st.session_state.edit_history = []
if 'analyzed_errors' not in st.session_state:
    st.session_state.analyzed_errors = []
if 'remove_noise_errors' not in st.session_state:
    st.session_state.remove_noise_errors = True  # Default enabled
if 'current_error_index' not in st.session_state:
    st.session_state.current_error_index = None
if 'last_removed_error' not in st.session_state:
    st.session_state.last_removed_error = None

def parse_xml_file_flexible(xml_content):
    """Parse XML file with multiple fallback strategies"""
    try:
        # DEBUG: Count occurrences of &#10; in the original XML
        debug_original_count = xml_content.count('&#10;')
        
        # PRESERVATION FIX: Replace &#10; with a placeholder that won't be converted to newlines
        # We'll use a unique placeholder that's unlikely to appear in the actual data
        placeholder = "__XML_LINE_BREAK_PLACEHOLDER__"
        preserved_xml = xml_content.replace('&#10;', placeholder)
        
        # Parse the XML while preserving the original structure
        root = ET.fromstring(preserved_xml)
        
        # Store original XML
        st.session_state.original_xml_content = xml_content
        
        # Parse header mappings first
        parse_header_mappings(root)
        
        # Define namespaces
        namespaces = {
            'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
            'o': 'urn:schemas-microsoft-com:office:office',
            'x': 'urn:schemas-microsoft-com:office:excel',
        }
        
        # Find worksheets
        worksheets = root.findall('.//ss:Worksheet', namespaces)
        if not worksheets:
            worksheets = root.findall('.//Worksheet')
        
        data = {}
        
        for ws in worksheets:
            # Get sheet name
            name = ws.get(f'{{{namespaces["ss"]}}}Name') if namespaces and 'ss' in namespaces else ws.get('ss:Name')
            if not name:
                name = ws.get('Name')
            if not name:
                name = f"Sheet_{len(data)+1}"
            
            # Find rows
            rows = ws.findall('.//ss:Row', namespaces)
            if not rows:
                rows = ws.findall('.//Row')
            
            if not rows:
                continue
            
            # Find header row
            headers = None
            header_row_index = -1
            
            for i, row in enumerate(rows):
                row_values = parse_row_cells(row, namespaces)
                
                # Check if this looks like a header row
                if not headers:
                    if any('EBELN' in str(cell).upper() for cell in row_values if cell):
                        headers = row_values
                        header_row_index = i
                        break
                    
                    # Also check for other common header patterns
                    for key in ['TYPE', 'MESSAGE', 'MATNR', 'LIFNR']:
                        if any(key in str(cell).upper() for cell in row_values if cell):
                            headers = row_values
                            header_row_index = i
                            break
                
                if headers:
                    break
            
            if not headers:
                continue
            
            # Extract data rows
            data_rows = []
            for i, row in enumerate(rows):
                if i <= header_row_index:
                    continue
                
                row_values = parse_row_cells(row, namespaces)
                
                # Create row dictionary with headers
                if any(row_values):  # Skip empty rows
                    row_dict = {}
                    for j, cell_value in enumerate(row_values):
                        if j < len(headers):
                            row_dict[headers[j]] = cell_value
                        else:
                            row_dict[f'Column{j+1}'] = cell_value
                    
                    if row_dict:
                        data_rows.append(row_dict)
            
            if data_rows:
                df = pd.DataFrame(data_rows)
                # Clean the dataframe
                df = df.replace('', pd.NA)
                df = df.dropna(how='all')
                df = df.dropna(axis=1, how='all')
                
                if not df.empty:
                    # Convert descriptive headers to technical names if needed
                    df = convert_to_technical_headers(df, name)
                    data[name] = df
        
        return data
    except Exception as e:
        st.error(f"Error parsing XML file: {str(e)}")
        return {}

def parse_row_cells(row, namespaces=None):
    """Parse cells from a row, handling sparse XML structure with ss:Index"""
    row_values = []
    expected_position = 0
    
    # Find cells
    if namespaces and 'ss' in namespaces:
        cells = row.findall('.//ss:Cell', namespaces)
    else:
        cells = row.findall('.//Cell')
    
    for cell in cells:
        # Get the column index from attributes
        col_index_attr = None
        
        # Try different ways to get the index attribute
        if namespaces and 'ss' in namespaces:
            # Try with namespace
            ns_attr = f'{{{namespaces["ss"]}}}Index'
            if ns_attr in cell.attrib:
                col_index_attr = cell.attrib[ns_attr]
        
        if not col_index_attr:
            # Try without namespace
            col_index_attr = cell.get('ss:Index') or cell.get('Index')
        
        if col_index_attr:
            # This cell has an explicit index (1-based)
            target_position = int(col_index_attr) - 1  # Convert to 0-based
            
            # Fill in missing cells between expected and target
            while len(row_values) < target_position:
                row_values.append("")
            
            # Update expected position to be AFTER this cell
            expected_position = target_position + 1
        else:
            # No explicit index - use expected position
            target_position = expected_position
            
            # Fill up to the expected position
            while len(row_values) < target_position:
                row_values.append("")
            
            # Move expected position for next cell
            expected_position += 1
        
        # Extract data from cell
        if namespaces and 'ss' in namespaces:
            data_elem = cell.find('.//ss:Data', namespaces)
        else:
            data_elem = cell.find('.//Data')
        
        cell_value = data_elem.text.strip() if data_elem is not None and data_elem.text else ""
        
        # PRESERVATION FIX: Replace our placeholder back with &#10; for display
        placeholder = "__XML_LINE_BREAK_PLACEHOLDER__"
        if placeholder in cell_value:
            cell_value = cell_value.replace(placeholder, '&#10;')
        
        # Ensure we have space at target_position
        while len(row_values) <= target_position:
            row_values.append("")
        
        # Place the value at the correct position
        row_values[target_position] = cell_value
    
    return row_values


def parse_header_mappings(root):
    """Parse the header mapping information from XML"""
    try:
        namespaces = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
        
        # Find all rows that contain header mappings
        for row in root.findall('.//Row'):
            cells = row.findall('.//Cell/Data', namespaces) or row.findall('.//ss:Cell/ss:Data', namespaces)
            
            if cells and len(cells) >= 7:
                description = cells[0].text.strip() if cells[0].text else ""
                sheet_name = cells[5].text.strip() if len(cells) > 5 and cells[5].text else ""
                technical_name = cells[6].text.strip() if len(cells) > 6 and cells[6].text else ""
                
                if description and technical_name and sheet_name:
                    if sheet_name not in st.session_state.header_mappings:
                        st.session_state.header_mappings[sheet_name] = {}
                    
                    # Store both directions
                    st.session_state.header_mappings[sheet_name][technical_name] = description
                    st.session_state.header_mappings[sheet_name][description] = technical_name
    except Exception as e:
        # Silently handle mapping parsing errors
        pass

def convert_to_technical_headers(df, sheet_name):
    """Convert descriptive headers to technical names"""
    if sheet_name in st.session_state.header_mappings:
        mapping = st.session_state.header_mappings[sheet_name]
        new_columns = []
        
        for col in df.columns:
            if col in mapping:
                # Check if this is a descriptive name that should map to technical name
                technical_name = mapping.get(col)
                if technical_name and technical_name in mapping and mapping[technical_name] == col:
                    new_columns.append(technical_name)
                else:
                    new_columns.append(col)
            else:
                new_columns.append(col)
        
        df.columns = new_columns
    
    return df

def convert_to_descriptive_headers(df, sheet_name):
    """Convert technical headers to descriptive names for display"""
    if sheet_name in st.session_state.header_mappings:
        mapping = st.session_state.header_mappings[sheet_name]
        descriptive_columns = []
        
        for col in df.columns:
            if col in mapping:
                descriptive_columns.append(mapping[col])
            else:
                descriptive_columns.append(col)
        
        display_df = df.copy()
        display_df.columns = descriptive_columns
        return display_df
    
    return df.copy()

def parse_excel_xlsx(uploaded_file):
    """Parse the Excel XLSX file and extract data from all worksheets"""
    try:
        # Reset file pointer
        uploaded_file.seek(0)
        workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        
        data = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Convert sheet to DataFrame
            data_list = []
            headers = None
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                # Skip empty rows
                if not any(cell is not None for cell in row):
                    continue
                    
                # Find header row - look for common header patterns
                if headers is None:
                    # Check if this row contains typical header values
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    
                    # Check for migration data headers (EBELN, etc.)
                    if any('EBELN' in str(val) for val in row_values):
                        headers = row_values
                        continue
                    
                    # Check for error log headers (Type, Message, etc.)
                    if any('Type' in str(val) or 'Message' in str(val) for val in row_values):
                        headers = row_values
                        continue
                
                if headers and i > sheet.min_row:  # Start collecting data after headers
                    row_dict = {}
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            if cell is not None:
                                # Handle different data types
                                if isinstance(cell, datetime):
                                    row_dict[headers[j]] = cell.strftime('%Y-%m-%d %H:%M:%S')
                                else:
                                    row_dict[headers[j]] = str(cell)
                            else:
                                row_dict[headers[j]] = ""
                    
                    if any(row_dict.values()):  # Only add non-empty rows
                        data_list.append(row_dict)
            
            if data_list:
                # Clean the data
                df = pd.DataFrame(data_list)
                # Remove completely empty rows
                df = df.dropna(how='all')
                # Remove completely empty columns
                df = df.dropna(axis=1, how='all')
                data[sheet_name] = df
        
        workbook.close()
        return data
    except Exception as e:
        st.error(f"Error parsing Excel file: {str(e)}")
        return {}

def parse_error_log_from_excel(error_data):
    """Extract error information from Excel data"""
    # Look for a sheet that contains error information
    error_df = None
    
    for sheet_name, df in error_data.items():
        # Check if this sheet contains error data
        if 'Type' in df.columns and 'Message Title' in df.columns:
            error_df = df
            break
    
    if error_df is None:
        # If no specific error sheet found, try to find error-like data
        for sheet_name, df in error_data.items():
            # Look for columns that might contain error information
            potential_cols = [col for col in df.columns if 'error' in col.lower() or 'message' in col.lower()]
            if potential_cols:
                error_df = df
                break
    
    return error_df

def find_related_records(data, error_df, error_index):
    """Find records related to a specific error"""
    error = error_df.iloc[error_index]
    related_records = {}
    
    # Extract PO number from error message if available
    po_number = None
    if "Source Record:" in str(error.get('Message Title', '')):
        match = re.search(r'Source Record: (\d+)', str(error.get('Message Title', '')))
        if match:
            po_number = match.group(1)
    
    # Find related records in all worksheets
    for sheet_name, df in data.items():
        if po_number and 'EBELN' in df.columns:
            matching_records = df[df['EBELN'] == po_number]
            if not matching_records.empty:
                related_records[sheet_name] = matching_records
    
    return related_records, po_number

def find_problematic_records(data, error_message, error_class, error_number):
    """Find records that might be causing the specific error"""
    problematic_records = {}
    
    # Extract material number from error message
    material_match = re.search(r'material (\d+)', error_message.lower())
    material_number = material_match.group(1) if material_match else None
    
    # Extract plant if mentioned
    plant_match = re.search(r'plant (\w+)', error_message.lower())
    plant_number = plant_match.group(1) if plant_match else None
    
    # Handle specific error patterns
    if material_number:
        # Try different formats of the material number (with/without leading zeros)
        material_variants = [
            material_number,  # Original format from error message
            material_number.zfill(10),  # Padded to 10 digits
            material_number.zfill(18),  # Padded to 18 digits
            material_number.lstrip('0'),  # Remove leading zeros
        ]
        
        # Remove duplicates while preserving order
        material_variants = list(dict.fromkeys(material_variants))
        
        # Search for the specific material in all relevant worksheets
        if 'Item Data' in data:
            df = data['Item Data']
            if 'MATNR' in df.columns:
                material_records = pd.DataFrame()  # Empty dataframe to collect results
                
                for variant in material_variants:
                    variant_records = df[df['MATNR'] == variant]
                    if not variant_records.empty:
                        material_records = pd.concat([material_records, variant_records])
                
                if not material_records.empty:
                    problematic_records['Item Data'] = material_records
        
        # Check Account Assignment for this material
        if 'Account Assignment' in data:
            df = data['Account Assignment']
            if 'EBELN' in df.columns and 'EBELP' in df.columns:
                # Find POs that contain this material
                if 'Item Data' in data and 'EBELN' in data['Item Data'].columns:
                    item_df = data['Item Data']
                    po_with_material = []
                    
                    for variant in material_variants:
                        variant_po = item_df[item_df['MATNR'] == variant]['EBELN'].unique()
                        po_with_material.extend(variant_po)
                    
                    # Remove duplicates
                    po_with_material = list(dict.fromkeys(po_with_material))
                    
                    if len(po_with_material) > 0:
                        # Find account assignments for these POs
                        acc_records = df[df['EBELN'].isin(po_with_material)]
                        if not acc_records.empty:
                            problematic_records['Account Assignment'] = acc_records
        
        # Check if this is an account assignment error
        if 'acc. ass. cat.' in error_message.lower():
            if 'Account Assignment' in data:
                df = data['Account Assignment']
                if 'KNTTP' in df.columns:  # Account Assignment Category
                    # Find records with empty KNTTP for this material
                    if 'EBELN' in df.columns and 'EBELP' in df.columns:
                        # Get POs with this material
                        if 'Item Data' in data:
                            item_df = data['Item Data']
                            po_with_material = []
                            
                            for variant in material_variants:
                                variant_po = item_df[item_df['MATNR'] == variant]['EBELN'].unique()
                                po_with_material.extend(variant_po)
                            
                            # Remove duplicates
                            po_with_material = list(dict.fromkeys(po_with_material))
                            
                            if len(po_with_material) > 0:
                                # Find account assignments without category
                                empty_knttp = df[
                                    (df['EBELN'].isin(po_with_material)) & 
                                    (df['KNTTP'].isna() | (df['KNTTP'] == ''))
                                ]
                                if not empty_knttp.empty:
                                    problematic_records['Account Assignment'] = empty_knttp
        
        # Check short text error
        if 'short text' in error_message.lower():
            if 'Item Data' in data:
                df = data['Item Data']
                if 'TXZ01' in df.columns:
                    # Find records with empty short text for this material
                    empty_text = pd.DataFrame()  # Empty dataframe to collect results
                    
                    for variant in material_variants:
                        variant_empty = df[
                            (df['MATNR'] == variant) & 
                            (df['TXZ01'].isna() | (df['TXZ01'] == ''))
                        ]
                        if not variant_empty.empty:
                            empty_text = pd.concat([empty_text, variant_empty])
                    
                    if not empty_text.empty:
                        problematic_records['Item Data'] = empty_text
        
        # Check plant maintenance error
        if plant_number and 'not maintained in plant' in error_message.lower():
            if 'Item Data' in data:
                df = data['Item Data']
                if 'MATNR' in df.columns and 'WERKS' in df.columns:
                    # Find records where material is not maintained in the specified plant
                    plant_issue = pd.DataFrame()  # Empty dataframe to collect results
                    
                    for variant in material_variants:
                        variant_issue = df[
                            (df['MATNR'] == variant) & 
                            (df['WERKS'] != plant_number)
                        ]
                        if not variant_issue.empty:
                            plant_issue = pd.concat([plant_issue, variant_issue])
                    
                    if not plant_issue.empty:
                        problematic_records['Item Data'] = plant_issue
        
        # Check source list error
        if 'source list' in error_message.lower():
            if 'Item Data' in data:
                df = data['Item Data']
                if 'MATNR' in df.columns:
                    # Just show all records for this material
                    material_records = pd.DataFrame()  # Empty dataframe to collect results
                    
                    for variant in material_variants:
                        variant_records = df[df['MATNR'] == variant]
                        if not variant_records.empty:
                            material_records = pd.concat([material_records, variant_records])
                    
                    if not material_records.empty:
                        problematic_records['Item Data'] = material_records
    
    # Handle other generic errors
    elif error_class == 'ME' and error_number == '83':  # "Enter Purch. Group"
        if 'Header Data' in data:
            df = data['Header Data']
            if 'EKGRP' in df.columns:
                empty_ekgrp = df[df['EKGRP'].isna() | (df['EKGRP'] == '')]
                if not empty_ekgrp.empty:
                    problematic_records['Header Data'] = empty_ekgrp
    
    elif 'delivery date' in error_message.lower():
        if 'Item Data' in data:
            df = data['Item Data']
            if 'LDATE' in df.columns:
                problematic_dates = df[df['LDATE'].isna() | (df['LDATE'] == '')]
                if not problematic_dates.empty:
                    problematic_records['Item Data'] = problematic_dates
        
        if 'Schedule Line' in data:
            df = data['Schedule Line']
            if 'EINDT' in df.columns:
                problematic_dates = df[df['EINDT'].isna() | (df['EINDT'] == '')]
                if not problematic_dates.empty:
                    problematic_records['Schedule Line'] = problematic_dates
    
    return problematic_records

def create_editable_dataframe(df, sheet_name, key_prefix, error_reference=None):
    """Create an editable dataframe widget with descriptive headers"""
    # Convert to descriptive headers for display
    display_df = convert_to_descriptive_headers(df, sheet_name)
    
    # Create a unique key for this editor instance
    # Include error_index in the key to differentiate between different errors
    editor_key = f"editor_{key_prefix}_{sheet_name}"
    
    # Store the original state to detect changes
    original_key = f'original_df_{editor_key}'
    
    # Check if we're viewing a different error than before
    # We need to track which error's data we're currently displaying
    error_tracking_key = f'current_error_{sheet_name}'
    
    # Reset stored data if we're showing a different error
    if error_tracking_key not in st.session_state:
        st.session_state[error_tracking_key] = key_prefix
    elif st.session_state[error_tracking_key] != key_prefix:
        # We're showing data for a different error - reset the stored data
        if original_key in st.session_state:
            del st.session_state[original_key]
        st.session_state[error_tracking_key] = key_prefix
    
    # Initialize or update the stored original
    if original_key not in st.session_state:
        # Store a fresh copy for this error
        st.session_state[original_key] = display_df.copy()
    
    # Get the stored original
    original_df = st.session_state[original_key]
    
    # Use Streamlit's data editor for editing
    edited_df = st.data_editor(
        display_df,
        key=editor_key,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True
    )
    
    # IMPORTANT: Only detect changes if this is the same error we initialized with
    # Check if the data shapes match (if not, we're looking at different data)
    same_data_structure = (
        len(original_df.columns) == len(display_df.columns) and
        list(original_df.columns) == list(display_df.columns) and
        len(original_df) == len(display_df)
    )
    
    # Also check if the actual content matches (first few rows as a sanity check)
    content_matches = True
    if same_data_structure and len(original_df) > 0:
        # Compare first row as a quick check
        for col in original_df.columns:
            if col in display_df.columns:
                orig_val = str(original_df.iloc[0][col]) if pd.notna(original_df.iloc[0][col]) else ""
                disp_val = str(display_df.iloc[0][col]) if pd.notna(display_df.iloc[0][col]) else ""
                if orig_val != disp_val:
                    content_matches = False
                    break
    
    # Only check for changes if we're looking at the same data we stored
    if same_data_structure and content_matches:
        # Check if there are changes
        if not edited_df.equals(original_df):
            # Find what changed
            changes = []
            
            # Check cell-by-cell changes
            for i in range(min(len(edited_df), len(original_df))):
                for j, col_name in enumerate(edited_df.columns):
                    new_value = edited_df.iloc[i, j]
                    old_value = original_df.iloc[i, j] if j < len(original_df.columns) else ""
                    
                    if pd.notna(new_value) and str(new_value) != str(old_value):
                        # Get the technical name
                        tech_name = get_technical_name(sheet_name, col_name)
                        
                        change = {
                            'sheet': sheet_name,
                            'column': tech_name,
                            'row': i,
                            'old_value': str(old_value),
                            'new_value': str(new_value),
                            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        }
                        
                        # Add error reference if provided
                        if error_reference and error_reference != "No specific reference":
                            change['error_reference'] = error_reference
                        
                        changes.append(change)
            
            # Check for new rows
            if len(edited_df) > len(original_df):
                for i in range(len(original_df), len(edited_df)):
                    for j, col_name in enumerate(edited_df.columns):
                        new_value = edited_df.iloc[i, j]
                        if pd.notna(new_value):
                            # Get the technical name
                            tech_name = get_technical_name(sheet_name, col_name)
                            
                            change = {
                                'sheet': sheet_name,
                                'column': tech_name,
                                'row': i,
                                'old_value': "",
                                'new_value': str(new_value),
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            
                            # Add error reference if provided
                            if error_reference and error_reference != "No specific reference":
                                change['error_reference'] = error_reference
                            
                            changes.append(change)
            
            # Update edit history with new changes only
            for change in changes:
                # Check if this exact edit already exists
                exists = False
                for existing in st.session_state.edit_history:
                    if (existing['sheet'] == change['sheet'] and 
                        existing['column'] == change['column'] and 
                        existing['row'] == change['row'] and 
                        existing['new_value'] == change['new_value']):
                        exists = True
                        break
                
                if not exists:
                    st.session_state.edit_history.append(change)
            
            # Update the stored original
            st.session_state[original_key] = edited_df.copy()
    else:
        # We're looking at different data (different error) - update the stored baseline
        st.session_state[original_key] = display_df.copy()
    
    return df


def get_technical_name(sheet_name, descriptive_name):
    """Get technical name from descriptive name"""
    if sheet_name in st.session_state.header_mappings:
        mapping = st.session_state.header_mappings[sheet_name]
        if descriptive_name in mapping:
            tech_name = mapping[descriptive_name]
            # Make sure we're returning the technical name
            if tech_name in mapping and mapping[tech_name] == descriptive_name:
                return tech_name
    return descriptive_name

def update_xml_with_changes(xml_content, modified_dataframes):
    """Update the original XML with the modified data"""
    try:
        # PRESERVATION FIX: Replace &#10; and &#13; with placeholders before parsing
        # This prevents ElementTree from converting them to actual newlines/carriage returns
        placeholder_10 = "__XML_LINE_BREAK_PLACEHOLDER__"
        placeholder_13 = "__XML_CARRIAGE_RETURN_PLACEHOLDER__"
        placeholder_quotes = "__XML_QUOTES_PLACEHOLDER__"
        preserved_xml = xml_content.replace('&#10;', placeholder_10).replace('&#13;', placeholder_13).replace('&quot;', placeholder_quotes)
                
        # Parse the XML while preserving the original structure
        root = ET.fromstring(preserved_xml)
        
        # Define namespaces exactly as they appear in the original
        namespaces = {
            'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
            'o': 'urn:schemas-microsoft-com:office:office',
            'x': 'urn:schemas-microsoft-com:office:excel',
            'dt': 'uuid:C2F41010-65B3-11d1-A29F-00AA00C14882',
            'html': 'http://www.w3.org/TR/REC-html40'
        }
        
        # Track changes made
        changes_made = []
        
        # Apply changes using the edit history
        if 'edit_history' in st.session_state and st.session_state.edit_history:
            # Group edits by sheet
            edits_by_sheet = {}
            for edit in st.session_state.edit_history:
                sheet_name = edit['sheet']
                if sheet_name not in edits_by_sheet:
                    edits_by_sheet[sheet_name] = []
                edits_by_sheet[sheet_name].append(edit)
            
            # Apply changes for each sheet
            for sheet_name, edits in edits_by_sheet.items():
                # Find the worksheet
                worksheet = None
                for child in root:
                    if 'Worksheet' in child.tag:
                        ws_name = child.get(f'{{{namespaces["ss"]}}}Name')
                        if ws_name == sheet_name:
                            worksheet = child
                            break
                
                if worksheet is None:
                    continue
                
                # Find rows in the worksheet
                rows = worksheet.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Row')
                
                if not rows:
                    continue
                
                # Find header row
                header_row_index = -1
                header_technical_names = []
                
                for i, row in enumerate(rows):
                    cells = row.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                    row_data = []
                    
                    for cell in cells:
                        data_elem = cell.find('.//{urn:schemas-microsoft-com:office:spreadsheet}Data')
                        if data_elem is not None and data_elem.text:
                            row_data.append(data_elem.text.strip())
                    
                    # Check if this is a header row
                    if any('EBELN' in cell.upper() for cell in row_data if cell):
                        header_technical_names = row_data
                        header_row_index = i
                        break
                
                if header_row_index == -1:
                    continue
                
                # Based on the debug output, the structure is:
                # Row 4: Header
                # Row 5: Technical info (ETE;80;0;C;80;0)
                # Row 6: Section headers (Key, General Data, etc.)
                # Row 7: Field descriptions
                # Row 8 onwards: Actual data
                
                # The actual data rows start after the description row
                # Let's find the first row that doesn't contain technical info or descriptions
                data_start_row = header_row_index + 1
                
                # Skip technical info row
                if data_start_row < len(rows):
                    # Check if this is the technical info row
                    row = rows[data_start_row]
                    cells = row.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                    row_data = []
                    
                    for cell in cells:
                        data_elem = cell.find('.//{urn:schemas-microsoft-com:office:spreadsheet}Data')
                        if data_elem is not None and data_elem.text:
                            row_data.append(data_elem.text.strip())
                    
                    # Check if this is technical info (all cells have semicolon-separated values)
                    is_tech_info = all(
                        ';' in cell and len(cell.split(';')) >= 5 
                        for cell in row_data if cell
                    )
                    
                    if is_tech_info:
                        data_start_row += 1
                
                # Skip section headers row
                if data_start_row < len(rows):
                    row = rows[data_start_row]
                    cells = row.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                    row_data = []
                    
                    for cell in cells:
                        data_elem = cell.find('.//{urn:schemas-microsoft-com:office:spreadsheet}Data')
                        if data_elem is not None and data_elem.text:
                            row_data.append(data_elem.text.strip())
                    
                    # Check if this is section headers (contains words like "Key", "General Data")
                    if any(word in ' '.join(row_data) for word in ['Key', 'General Data', 'Delivery', 'Incoterms']):
                        data_start_row += 1
                
                # Skip field descriptions row
                if data_start_row < len(rows):
                    row = rows[data_start_row]
                    cells = row.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                    row_data = []
                    
                    for cell in cells:
                        data_elem = cell.find('.//{urn:schemas-microsoft-com:office:spreadsheet}Data')
                        if data_elem is not None and data_elem.text:
                            row_data.append(data_elem.text.strip())
                    
                    # Check if this is field descriptions (contains long text with newlines)
                    if any('\n\n' in cell for cell in row_data if cell):
                        data_start_row += 1
                
                # Create a map of data rows by their EBELN (PO number) for easier lookup
                data_rows_by_ebeln = {}
                data_rows_by_index = {}  # Also map by dataframe index
                
                for i in range(data_start_row, len(rows)):
                    row = rows[i]
                    cells = row.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                    row_data = []
                    
                    current_position = 0
                    for cell in cells:
                        # Get cell position from ss:Index or use sequential
                        index_attr = cell.get(f'{{{namespaces["ss"]}}}Index')
                        if index_attr:
                            cell_position = int(index_attr) - 1  # Convert to 0-based
                        else:
                            cell_position = current_position
                        
                        data_elem = cell.find('.//{urn:schemas-microsoft-com:office:spreadsheet}Data')
                        if data_elem is not None and data_elem.text:
                            # PRESERVATION FIX: Replace placeholders back with actual newlines/carriage returns
                            # This ensures that the data in the parsed XML has actual newlines/carriage returns
                            cell_text = data_elem.text.replace(placeholder_10, '\n').replace(placeholder_13, '\r').replace(placeholder_quotes, '"')
                            
                            # Ensure list is long enough
                            while len(row_data) <= cell_position:
                                row_data.append("")
                            row_data[cell_position] = cell_text.strip()
                        else:
                            # Ensure list is long enough
                            while len(row_data) <= cell_position:
                                row_data.append("")
                            row_data[cell_position] = ""
                        
                        current_position = cell_position + 1
                    
                    # If this row has an EBELN, add it to our map
                    if len(row_data) > 0 and row_data[0]:
                        data_rows_by_ebeln[row_data[0]] = (i, row, cells)
                        
                        # Also map by dataframe index (assuming the order matches)
                        df_index = i - data_start_row
                        data_rows_by_index[df_index] = (i, row, cells)
                
                # Apply each edit
                for edit in edits:
                    column_name = edit['column']
                    row_index = edit['row']  # This is the index in the dataframe (0-based)
                    new_value = edit['new_value']
                    
                    # Find the column index in the header
                    if column_name in header_technical_names:
                        col_index = header_technical_names.index(column_name)
                        
                        # Try to find the row by matching the EBELN from the edit
                        # First, get the EBELN from the edit if available
                        edit_ebeln = None
                        if 'EBELN' in edit:
                            edit_ebeln = edit['EBELN']
                        
                        # Also try to get the EBELN from the modified dataframes if available
                        if not edit_ebeln and sheet_name in modified_dataframes:
                            for df_key, modified_df in modified_dataframes[sheet_name].items():
                                if row_index < len(modified_df) and 'EBELN' in modified_df.columns:
                                    edit_ebeln = modified_df.iloc[row_index]['EBELN']
                                    break
                        
                        if edit_ebeln and edit_ebeln in data_rows_by_ebeln:
                            xml_row_index, row, cells = data_rows_by_ebeln[edit_ebeln]
                        elif row_index in data_rows_by_index:
                            xml_row_index, row, cells = data_rows_by_index[row_index]
                        else:
                            # Fallback to the original calculation
                            xml_row_index = data_start_row + row_index
                            
                            if xml_row_index < len(rows):
                                row = rows[xml_row_index]
                                cells = row.findall('.//{urn:schemas-microsoft-com:office:spreadsheet}Cell')
                            else:
                                continue
                        
                        if col_index < len(cells):
                            # Find the correct cell accounting for ss:Index
                            target_cell = None
                            current_position = 0
                            
                            for cell in cells:
                                # Check for ss:Index attribute
                                index_attr = cell.get(f'{{{namespaces["ss"]}}}Index')
                                if index_attr:
                                    cell_position = int(index_attr) - 1  # Convert to 0-based
                                else:
                                    cell_position = current_position
                                
                                if cell_position == col_index:
                                    target_cell = cell
                                    break
                                
                                current_position = cell_position + 1
                            
                            cell = target_cell
                        
                        if cell:
                            # Find or create Data element
                            data_elem = cell.find('.//{urn:schemas-microsoft-com:office:spreadsheet}Data')
                            if data_elem is None:
                                data_elem = ET.SubElement(cell, f'{{{namespaces["ss"]}}}Data')
                                data_elem.set(f'{{{namespaces["ss"]}}}Type', 'String')
                            
                            old_value = data_elem.text if data_elem.text else ""
                            
                            if str(new_value) != str(old_value):                                
                                # PRESERVATION FIX: Keep the new value as is (with actual newlines/carriage returns)
                                # We'll convert back to &#10; and &#13; in the final step
                                data_elem.text = str(new_value)
                                
                                changes_made.append({
                                    'sheet': sheet_name,
                                    'column': column_name,
                                    'row': row_index,
                                    'old_value': old_value,
                                    'new_value': new_value
                                })
        
        # Register namespaces to maintain formatting
        for prefix, uri in namespaces.items():
            ET.register_namespace(prefix, uri)
        
        # Convert to string
        xml_str = ET.tostring(root, encoding='unicode')
                
        # PRESERVATION FIX: Replace the placeholders back with &#10; and &#13; in the final XML
        # This ensures that all original entities and new newlines/carriage returns are properly encoded
        xml_str = xml_str.replace(placeholder_10, '&#10;').replace(placeholder_13, '&#13;').replace(placeholder_quotes, '&quot;')
        
        # PRESERVATION FIX: Also replace any actual newlines/carriage returns in Data elements with &#10;/&#13;
        # This handles newlines/carriage returns that were added during editing
        def fix_newlines_in_data(match):
            opening_tag = match.group(1)
            content = match.group(2)
            closing_tag = match.group(3)
            
            # Replace newlines with &#10; and carriage returns with &#13; in the content
            content = content.replace('\n', '&#10;').replace('\r', '&#13;').replace('"', '&quot;')
            
            return opening_tag + content + closing_tag
        
        # Pattern to match Data elements with content
        import re
        data_content_pattern = r'(<Data[^>]*>)([^<]*)(</Data>)'
        xml_str = re.sub(data_content_pattern, fix_newlines_in_data, xml_str)
        
        # Add XML declaration and processing instruction
        xml_declaration = '<?xml version="1.0"?>\n<?mso-application progid="Excel.Sheet"?>\n'
        
        # Use a custom function to fix the namespace structure and formatting
        def fix_xml_formatting(xml_str):
            # PRESERVATION FIX: Store the current count of entities before processing
            before_count_10 = xml_str.count('&#10;')
            before_count_13 = xml_str.count('&#13;')
			    
            
            # Parse the XML with minidom for better control
            try:
                # PRESERVATION FIX: Replace &#10; and &#13; with temporary placeholders before minidom processing
                # This prevents minidom from converting them to actual newlines/carriage returns
                temp_placeholder_10 = "__TEMP_LINE_BREAK__"
                temp_placeholder_13 = "__TEMP_CARRIAGE_RETURN__"
                temp_placeholder_quotes = "__TEMP_QUOTES_RETURN__"
                temp_xml = xml_str.replace('&#10;', temp_placeholder_10).replace('&#13;', temp_placeholder_13).replace('&quot;', temp_placeholder_quotes)
                
                dom = minidom.parseString(temp_xml)
                pretty_xml = dom.toprettyxml(indent="\t")
                
                # Remove the XML declaration added by minidom
                pretty_xml = pretty_xml.replace('<?xml version="1.0" ?>', '')
                
                # PRESERVATION FIX: Replace the temporary placeholders back with &#10; and &#13;
                pretty_xml = pretty_xml.replace(temp_placeholder_10, '&#10;').replace(temp_placeholder_13, '&#13;').replace(temp_placeholder_quotes, '&quot;')
            except:
                # If minidom fails, use the original string
                pretty_xml = xml_str
            
            # Fix the Workbook element to match the original format
            import re
            
            # Pattern to match the Workbook element
            workbook_pattern = r'<ss:Workbook[^>]*>'
            
            # The replacement should match the original format exactly
            workbook_replacement = '<Workbook\n\txmlns="urn:schemas-microsoft-com:office:spreadsheet"\n\txmlns:o="urn:schemas-microsoft-com:office:office"\n\txmlns:x="urn:schemas-microsoft-com:office:excel"\n\txmlns:dt="uuid:C2F41010-65B3-11d1-A29F-00AA00C14882"\n\txmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"\n\txmlns:html="http://www.w3.org/TR/REC-html40">'
            
            # Replace the Workbook element
            pretty_xml = re.sub(workbook_pattern, workbook_replacement, pretty_xml)
            
            # Fix the DocumentProperties element to match the original format
            doc_props_pattern = r'<o:DocumentProperties[^>]*>'
            doc_props_replacement = '\t<DocumentProperties\n\t\txmlns="urn:schemas-microsoft-com:office:office">'
            pretty_xml = re.sub(doc_props_pattern, doc_props_replacement, pretty_xml)
            
            # Fix the CustomDocumentProperties element to match the original format
            custom_doc_props_pattern = r'<o:CustomDocumentProperties[^>]*>'
            custom_doc_props_replacement = '\t<CustomDocumentProperties\n\t\txmlns="urn:schemas-microsoft-com:office:office">'
            pretty_xml = re.sub(custom_doc_props_pattern, custom_doc_props_replacement, pretty_xml)
            
            # Fix OfficeDocumentSettings element to match the original format
            office_doc_settings_pattern = r'<o:OfficeDocumentSettings[^>]*>'
            office_doc_settings_replacement = '\t<OfficeDocumentSettings\n\t\txmlns="urn:schemas-microsoft-com:office:office">'
            pretty_xml = re.sub(office_doc_settings_pattern, office_doc_settings_replacement, pretty_xml)
            
            # Fix ExcelWorkbook element to match the original format
            excel_workbook_pattern = r'<x:ExcelWorkbook[^>]*>'
            excel_workbook_replacement = '\t<ExcelWorkbook\n\t\txmlns="urn:schemas-microsoft-com:office:excel">'
            pretty_xml = re.sub(excel_workbook_pattern, excel_workbook_replacement, pretty_xml)
            
            # Fix WorksheetOptions element to match the original format
            worksheet_options_pattern = r'<x:WorksheetOptions[^>]*>'
            worksheet_options_replacement = '\t\t<WorksheetOptions\n\t\t\txmlns="urn:schemas-microsoft-com:office:excel">'
            pretty_xml = re.sub(worksheet_options_pattern, worksheet_options_replacement, pretty_xml)
            
            # Fix the closing tags
            pretty_xml = pretty_xml.replace('</o:DocumentProperties>', '\t</DocumentProperties>')
            pretty_xml = pretty_xml.replace('</o:CustomDocumentProperties>', '\t</CustomDocumentProperties>')
            pretty_xml = pretty_xml.replace('</o:OfficeDocumentSettings>', '\t</OfficeDocumentSettings>')
            pretty_xml = pretty_xml.replace('</x:ExcelWorkbook>', '\t</ExcelWorkbook>')
            pretty_xml = pretty_xml.replace('</x:WorksheetOptions>', '\t\t</WorksheetOptions>')
            pretty_xml = pretty_xml.replace('</ss:Workbook>', '</Workbook>')
            
            # Fix indentation for DocumentProperties
            pretty_xml = pretty_xml.replace('\t\t<DocumentProperties', '\t<DocumentProperties')
            pretty_xml = pretty_xml.replace('\t\t</DocumentProperties>', '\t</DocumentProperties>')
            
            # Fix indentation for CustomDocumentProperties
            pretty_xml = pretty_xml.replace('\t\t<CustomDocumentProperties', '\t<CustomDocumentProperties')
            pretty_xml = pretty_xml.replace('\t\t</CustomDocumentProperties>', '\t</CustomDocumentProperties>')
            
            # Fix indentation for OfficeDocumentSettings
            pretty_xml = pretty_xml.replace('\t\t<OfficeDocumentSettings', '\t<OfficeDocumentSettings')
            pretty_xml = pretty_xml.replace('\t\t</OfficeDocumentSettings>', '\t</OfficeDocumentSettings>')
            
            # Fix indentation for ExcelWorkbook
            pretty_xml = pretty_xml.replace('\t\t<ExcelWorkbook', '\t<ExcelWorkbook')
            pretty_xml = pretty_xml.replace('\t\t</ExcelWorkbook>', '\t</ExcelWorkbook>')
            
            # Fix indentation for WorksheetOptions
            pretty_xml = pretty_xml.replace('\t\t\t<WorksheetOptions', '\t\t<WorksheetOptions')
            pretty_xml = pretty_xml.replace('\t\t\t</WorksheetOptions>', '\t\t</WorksheetOptions>')
            
            
            # PRESERVATION FIX: We don't need to process Data elements here anymore
            # since we've already handled all newlines and carriage returns above
            
            # Remove all namespace prefixes from elements, but preserve them in attributes
            lines = pretty_xml.split('\n')
            new_lines = []
            
            for line in lines:
                # Skip empty lines
                if not line.strip():
                    continue
                
                # Process each line to remove namespace prefixes from elements
                new_line = re.sub(r'<(/?)([a-z]+):', r'<\1', line)
                new_lines.append(new_line)
            
            return '\n'.join(new_lines)
        
        # Fix the XML formatting
        fixed_xml = fix_xml_formatting(xml_str)
        
        # FINAL DIRECT FIX: Replace all self-closing Data elements with explicit closing tags
        # Pattern to match self-closing Data elements with any attributes
        data_patterns = [
            r'<Data([^>]*)\s*/>',  # Standard case with space before />
            r'<Data([^>]*)/>',     # Case without space before />
        ]
        
        for pattern in data_patterns:
            # Replace with opening and closing tags
            fixed_xml = re.sub(pattern, r'<Data\1></Data>', fixed_xml)
        
        # Add our custom XML declaration
        result_xml = xml_declaration + fixed_xml
        
        return result_xml, changes_made
        
    except Exception as e:
        st.error(f"Error updating XML: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return None, []



def extract_error_reference(error_message):
    """Extract meaningful references from error messages"""
    if not isinstance(error_message, str):
        return "No reference"
    
    # Common patterns for extracting references
    patterns = [
        # PO Number patterns
        (r'Source Record:\s*(\d+)', 'EBELN'),
        (r'PO\s*(\d+)', 'EBELN'),
        (r'purchase order\s*(\d+)', 'EBELN', True),
        (r'Purch\.\s*Doc\.\s*(\d+)', 'EBELN'),
        
        # Material patterns
        (r'material\s*(\d+)', 'MATNR', True),
        (r'Material\s*(\d+)', 'MATNR'),
        (r'MATNR:\s*(\d+)', 'MATNR'),
        
        # Vendor patterns
        (r'vendor\s*(\d+)', 'LIFNR', True),
        (r'Vendor\s*(\d+)', 'LIFNR'),
        (r'LIFNR:\s*(\d+)', 'LIFNR'),
        
        # Plant patterns
        (r'plant\s*(\w+)', 'WERKS', True),
        (r'Plant\s*(\w+)', 'WERKS'),
        
        # Purchasing Group patterns
        (r'purchasing group\s*(\w+)', 'EKGRP', True),
        (r'Purchasing Group\s*(\w+)', 'EKGRP'),
        
        # Account Assignment patterns
        (r'account assignment\s*(\w+)', 'KNTTP', True),
        (r'Acc\. Ass\. Cat\.\s*(\w+)', 'KNTTP'),
    ]
    
    for pattern, field_name, *flags in patterns:
        match = re.search(pattern, error_message, re.IGNORECASE if len(flags) > 0 and flags[0] else 0)
        if match:
            value = match.group(1)
            return f"{field_name}: {value}"
    
    # If no specific pattern found, extract any number that might be a reference
    number_match = re.search(r'\b(\d{6,})\b', error_message)
    if number_match:
        # Check if it looks like a PO number (10 digits) or material number (18 digits)
        value = number_match.group(1)
        if len(value) == 10:
            return f"EBELN: {value}"
        elif len(value) == 18:
            return f"MATNR: {value}"
        else:
            return f"Reference: {value}"
    
    # Try to extract any alphanumeric code
    code_match = re.search(r'\b([A-Z0-9]{6,})\b', error_message)
    if code_match:
        return f"Code: {code_match.group(1)}"
    
    return "No specific reference"

def get_error_context(edit):
    """Get error context for an edit based on the sheet and column"""
    # This function would need access to the current error being analyzed
    # For now, we'll return a generic context
    return "Error context not available"

def create_edit_summary():
    """Create a summary of all edits made with delete buttons"""
    if not st.session_state.edit_history:
        st.info("No edits have been made yet.")
        return pd.DataFrame()
    
    st.markdown("### 📋 Edit Summary")
    
    # Initialize session state
    if 'confirm_delete_index' not in st.session_state:
        st.session_state.confirm_delete_index = None
    
    # Display each edit with a delete button
    for i, edit in enumerate(st.session_state.edit_history):
        # Get descriptive column name
        sheet_name = edit['sheet']
        tech_name = edit['column']
        descriptive_name = tech_name
        
        if sheet_name in st.session_state.header_mappings:
            mapping = st.session_state.header_mappings[sheet_name]
            if tech_name in mapping:
                descriptive_name = mapping[tech_name]
        
        error_reference = edit.get('error_reference', '')
        
        # Create a container for each edit
        with st.container():
            # Create columns layout
            cols = st.columns([0.5, 2, 2, 1, 3, 3, 3, 3])
            
            with cols[0]:
                # Show either delete button or confirmation
                if st.session_state.confirm_delete_index == i:
                    # Already clicked - show confirmation
                    st.write("🗑️")  # Icon placeholder
                else:
                    # Regular delete button
                    if st.button("🗑️", key=f"del_init_{i}"):
                        st.session_state.confirm_delete_index = i
                        st.rerun()
            
            with cols[1]:
                st.write(sheet_name)
            
            with cols[2]:
                st.write(descriptive_name)
            
            with cols[3]:
                st.write(edit['row'])
            
            with cols[4]:
                old_val = str(edit['old_value'])
                display_old = old_val[:30] + "..." if len(old_val) > 30 else old_val
                st.write(display_old)
            
            with cols[5]:
                new_val = str(edit['new_value'])
                display_new = new_val[:30] + "..." if len(new_val) > 30 else new_val
                st.write(display_new)
            
            with cols[6]:
                display_ref = str(error_reference)[:30] + "..." if len(str(error_reference)) > 30 else str(error_reference)
                st.write(display_ref)
            
            with cols[7]:
                st.write(edit['timestamp'])
            
            # Show confirmation section if this is the edit to delete
            if st.session_state.confirm_delete_index == i:
                with st.container():
                    st.markdown("<hr>", unsafe_allow_html=True)
                    
                    # Confirmation box
                    with st.container():
                        st.warning(f"⚠️ **Confirm deletion of this edit?**")
                        st.write(f"**{sheet_name}** - **{descriptive_name}** (Row {edit['row']})")
                        
                        col1, col2, col3, col4 = st.columns([1, 1, 1, 2])
                        
                        with col1:
                            if st.button("✅ Yes", key=f"confirm_yes_{i}", type="primary"):
                                # Delete the edit
                                deleted_edit = st.session_state.edit_history.pop(i)
                                
                                # Clear modified dataframes
                                st.session_state.modified_dataframes = {}
                                
                                # Reset confirmation
                                st.session_state.confirm_delete_index = None
                                
                                # Show success
                                st.success(f"✅ Deleted edit")
                                st.rerun()
                        
                        with col2:
                            if st.button("❌ No", key=f"confirm_no_{i}"):
                                st.session_state.confirm_delete_index = None
                                st.rerun()
                        
                        with col4:
                            st.caption("This action cannot be undone.")
                    
                    st.markdown("<hr>", unsafe_allow_html=True)
            
            st.markdown("---")

    # Add a "Clear All" button with confirmation
    if st.session_state.edit_history:
        st.markdown("---")
        
        # Check if we're in clear all confirmation mode
        if 'clear_all_confirmation' not in st.session_state:
            st.session_state.clear_all_confirmation = False
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            if not st.session_state.clear_all_confirmation:
                if st.button("🗑️ Clear All Edits", type="primary", key="clear_all_init"):
                    st.session_state.clear_all_confirmation = True
                    st.rerun()
            else:
                st.warning("Are you sure you want to delete ALL edits?")
                confirm_col1, confirm_col2, confirm_col3 = st.columns([1, 1, 1])
                
                with confirm_col1:
                    if st.button("✅ Yes, Delete All", type="primary"):
                        # Store count for message
                        edit_count = len(st.session_state.edit_history)
                        
                        # Clear everything
                        st.session_state.edit_history = []
                        st.session_state.modified_dataframes = {}
                        st.session_state.delete_confirmation_index = None
                        st.session_state.clear_all_confirmation = False
                        
                        # Show success
                        st.success(f"✅ Cleared all {edit_count} edits")
                        st.rerun()
                
                with confirm_col3:
                    if st.button("❌ Cancel"):
                        st.session_state.clear_all_confirmation = False
                        st.rerun()
    
    return pd.DataFrame(st.session_state.edit_history)

def is_noise_error(error_message):
    """Check if an error message is noise (doesn't contain any numbers)"""
    if not isinstance(error_message, str):
        return True
    # Check if there's at least one digit in the message
    return not any(char.isdigit() for char in error_message)

def filter_noise_errors(error_df):
    """Filter out noise errors that don't contain numbers in the message"""
    if 'Message Title' in error_df.columns:
        # Keep errors that have at least one digit in the message
        filtered_df = error_df[error_df['Message Title'].apply(lambda x: any(char.isdigit() for char in str(x)))]
        return filtered_df
    return error_df

def filter_out_analyzed_errors(error_df):
    """Filter out errors that have been marked as analyzed"""
    if not st.session_state.analyzed_errors:
        return error_df
    
    # Create a set of analyzed error indices for quick lookup
    analyzed_indices = {error['index'] for error in st.session_state.analyzed_errors}
    
    # Filter out rows that are in the analyzed errors
    filtered_df = error_df[~error_df.index.isin(analyzed_indices)].copy()
    return filtered_df

def mark_error_as_analyzed(error_index, error_info, error_context=None):
    """Mark an error as analyzed and add it to the session state"""
    # Check if this error is already in the analyzed list
    for error in st.session_state.analyzed_errors:
        if error['index'] == error_index:
            return
    
    # Extract error reference from message
    error_message = error_info.get('Message Title', '')
    error_reference = extract_error_reference(error_message)
    
    # Add the error to analyzed list
    st.session_state.analyzed_errors.append({
        'index': error_index,
        'message': error_message,
        'type': error_info.get('Type', ''),
        'class': error_info.get('Message Class', ''),
        'number': error_info.get('Message Number', ''),
        'reference': error_reference,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })
    
    # REMOVED: Don't update existing edit history here
    # This was causing false edit detections
    # The error_reference should only be added when making actual edits
    # not when marking errors as analyzed

def main():
    st.title("📊 Purchase Order Migration Error Analysis & Editor")
    
    # Create sidebar for settings
    with st.sidebar:
        st.header("⚙️ Settings")
        
        # Noise removal checkbox
        st.subheader("Noise Filtering")
        remove_noise = st.checkbox(
            "Remove errors without numbers",
            value=st.session_state.remove_noise_errors,
            help="Hide errors that don't contain any numbers in their description (usually noise)"
        )
        st.session_state.remove_noise_errors = remove_noise
        
        # Analyzed errors section
        st.subheader("Analyzed Errors")
        if st.session_state.analyzed_errors:
            st.write(f"**Analyzed errors:** {len(st.session_state.analyzed_errors)}")
            
            # Display analyzed errors with remove buttons
            for i, error in enumerate(st.session_state.analyzed_errors):
                col1, col2 = st.columns([3, 1])
                with col1:
                    # Display error type and reference
                    error_type = error['type']
                    error_ref = error['reference']
                    
                    # Truncate message if too long
                    error_msg = error['message']
                    if len(error_msg) > 30:
                        error_msg = error_msg[:27] + "..."
                    
                    st.write(f"**{error_type}** - {error_ref}")
                
                with col2:
                    if st.button("❌", key=f"remove_analyzed_{i}"):
                        # Store the removed error index to potentially reselect it
                        st.session_state.last_removed_error = error['index']
                        st.session_state.analyzed_errors.pop(i)
                        # Don't reset current error index - let it persist
                        st.rerun()
            
            # Clear all analyzed errors button
            if st.button("🗑️ Clear All Analyzed", type="secondary"):
                st.session_state.analyzed_errors = []
                st.rerun()
        else:
            st.write("No errors analyzed yet")
        
        st.markdown("---")
        st.subheader("Filter Options")
        
        # Initialize session state for filter if not exists
        if 'message_filter' not in st.session_state:
            st.session_state.message_filter = 'Error'
        
        # Create radio buttons for filter
        filter_options = ['Error', 'Warning', 'Success', 'All']
        selected_filter = st.radio(
            "Show Message Types:",
            options=filter_options,
            index=filter_options.index(st.session_state.message_filter),
            key="message_filter_radio"
        )
        
        # Update session state
        st.session_state.message_filter = selected_filter
        
        st.markdown("---")
        st.info("""
        **Filter Help:**
        - **Error**: Critical issues that need fixing
        - **Warning**: Potential issues to review
        - **Success**: Completed migrations
        - **All**: Show all message types
        """)
    
    # Add instructions
    with st.expander("📋 Instructions"):
        st.markdown("""
        ### Steps to analyze and fix migration errors:
        1. **Upload the migration data file** (XML or XLSX format)
        2. **Upload the error log file** (XLSX format)
        3. **Select an error** from the list to view related records
        4. **Edit the data** directly in the tables below
        5. **Download the modified XML** with all changes applied
        
        ### Features:
        - 🔍 **Smart error analysis** with pattern matching
        - ✏️ **In-place editing** of related records
        - 📋 **Descriptive headers** with technical name tooltips
        - 📊 **Edit summary** showing all changes made
        - 💾 **Download modified XML** with corrections applied
        
        ### Column Display:
        - Columns show **descriptive names** (e.g., "Item Number of Purchasing Document")
        - Hover over column headers to see **technical names** (e.g., "EBELP")
        """)
    
    # File upload sections
    st.header("📁 Upload Files")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Upload Migration Data File**")
        st.write("Supported: XML (.xml) or Excel (.xlsx)")
        data_file = st.file_uploader(
            "Choose migration data file", 
            type=['xml', 'xlsx'],
            key="data_file"
        )
        if data_file:
            st.success(f"✅ File uploaded: {data_file.name}")
    
    with col2:
        st.write("**Upload Error Log File**")
        st.write("Supported: Excel (.xlsx)")
        error_file = st.file_uploader(
            "Choose error log file", 
            type=['xlsx'],
            key="error_file"
        )
        if error_file:
            st.success(f"✅ File uploaded: {error_file.name}")
    
    if data_file and error_file:
        # Parse files
        with st.spinner("🔄 Parsing files..."):
            try:
                # Parse migration data file
                if data_file.type == 'text/xml':
                    xml_content = data_file.read().decode('utf-8')
                    data = parse_xml_file_flexible(xml_content)
                elif data_file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    data = parse_excel_xlsx(data_file)
                else:
                    st.error(f"❌ Unsupported file type: {data_file.type}")
                    return
                
                # Parse error log file (always Excel now)
                error_data = parse_excel_xlsx(error_file)
                error_df = parse_error_log_from_excel(error_data)
                
                if error_df is None:
                    st.error("❌ Could not find error data in the uploaded file")
                    st.write("Please ensure the error log file contains columns like 'Type' and 'Message Title'")
                    return
                
            except Exception as e:
                st.error(f"❌ Error parsing files: {str(e)}")
                return
        
        # Display error summary
        st.header("🚨 Error Summary")
        
        # Error statistics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_errors = len(error_df)
            st.metric("Total Messages", total_errors)
        
        with col2:
            error_count = len(error_df[error_df['Type'] == 'Error']) if 'Type' in error_df.columns else 0
            st.metric("Errors", error_count)
        
        with col3:
            warning_count = len(error_df[error_df['Type'] == 'Warning']) if 'Type' in error_df.columns else 0
            st.metric("Warnings", warning_count)
        
        with col4:
            success_count = len(error_df[error_df['Type'] == 'Success']) if 'Type' in error_df.columns else 0
            st.metric("Success", success_count)
        
        # Filter errors based on sidebar selection
        if 'Type' in error_df.columns:
            if st.session_state.message_filter != 'All':
                filtered_errors = error_df[error_df['Type'] == st.session_state.message_filter]
            else:
                filtered_errors = error_df
        else:
            filtered_errors = error_df
        
        # Apply noise filtering if enabled and showing Errors
        if st.session_state.remove_noise_errors and st.session_state.message_filter == 'Error':
            filtered_errors = filter_noise_errors(filtered_errors)
        
        # Filter out analyzed errors
        filtered_errors = filter_out_analyzed_errors(filtered_errors)
        
        # Error details section
        st.header("🔍 Error Details & Analysis")
        
        # Select an error to analyze
        error_options = []
        for i, (_, row) in enumerate(filtered_errors.iterrows()):
            msg_title = str(row.get('Message Title', row.get('Message', '')))[:80]
            error_type = str(row.get('Type', 'Unknown'))
            error_options.append(f"🔴 {i}: {error_type} - {msg_title}...")
        
        # Check if we should select a specific error (last removed or previous selection)
        default_index = 0
        if error_options:
            # If we just removed an analyzed error, try to find it in the current filtered list
            if st.session_state.last_removed_error is not None:
                # Check if the removed error is in the current filtered list
                if st.session_state.last_removed_error in filtered_errors.index:
                    # Find its position in the filtered list
                    try:
                        default_index = filtered_errors.index.get_loc(st.session_state.last_removed_error)
                        st.session_state.last_removed_error = None  # Reset after using
                    except:
                        pass
            # Otherwise, use the previously selected index if available and valid
            elif (st.session_state.current_error_index is not None and 
                  st.session_state.current_error_index < len(error_options)):
                default_index = st.session_state.current_error_index
        
        if error_options:
            selected_error = st.selectbox(
                "Select an Error to Analyze", 
                error_options,
                index=default_index,
                key="error_selector"
            )
            error_index = int(selected_error.split(":")[0].replace("🔴 ", ""))
            
            # Store the current error index for persistence
            st.session_state.current_error_index = error_index
            
            # Get the actual error
            # We need to get the error from the filtered dataframe
            error_row = filtered_errors.iloc[error_index]
            original_index = filtered_errors.index[error_index]  # Get the original index
            
            # Extract error reference for display
            error_message = error_row.get('Message Title', '')
            error_reference = extract_error_reference(error_message)
            
            # Display error details
            st.markdown("### 📝 Error Information")
            
            # Show error reference prominently
            if error_reference != "No specific reference":
                st.info(f"🔍 **Reference:** {error_reference}")
            
            # Add checkbox to mark as analyzed
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write(f"**Type:** {error_row.get('Type', 'N/A')}")
                st.write(f"**Message Class:** {error_row.get('Message Class', 'N/A')}")
                st.write(f"**Message Number:** {error_row.get('Message Number', 'N/A')}")
            
            with col2:
                st.write(f"**Message:** {error_row.get('Message Title', 'N/A')}")
                st.write(f"**Date/Time:** {error_row.get('Date and Time (UTC)', 'N/A')}")
            
            # Mark as analyzed checkbox
            analyzed_col1, analyzed_col2 = st.columns([4, 1])
            with analyzed_col2:
                if st.button("✅ Mark as Analyzed", key="mark_analyzed"):
                    # Store error context to link with edits
                    error_context = {
                        'sheet': 'Item Data',  # Default, could be determined from analysis
                        'reference': error_reference
                    }
                    mark_error_as_analyzed(original_index, error_row, error_context)
                    st.session_state.current_error_index = None  # Reset selection
                    st.rerun()
            
            # Find and display related records
            st.markdown("### 🔗 Analysis Results")
            
            # First try to find records by PO number
            related_records, po_number = find_related_records(data, filtered_errors, error_index)
            
            if po_number:
                st.info(f"📋 Purchase Order Number: **{po_number}**")
            
            # If no records found by PO number, try to find problematic records
            if not related_records:
                error_class = error_row.get('Message Class', '')
                error_number = error_row.get('Message Number', '')
                error_message = error_row.get('Message Title', '')
                
                problematic_records = find_problematic_records(data, error_message, error_class, error_number)
                
                if problematic_records:
                    related_records = problematic_records
            
            # Display all related records with editing capability
            if related_records:
                st.markdown("### 📊 Related Records - Edit Mode")
                
                # Store error reference for this editing session
                current_error_ref = error_reference
                
                for sheet_name, records in related_records.items():
                    with st.container():
                        st.markdown(f"#### 📋 {sheet_name}")
                        
                        # Create editable dataframe - pass error reference as parameter
                        edited_df = create_editable_dataframe(records, sheet_name, f"error_{error_index}", current_error_ref)
                        
                        # Add separator
                        st.markdown("---")
            else:
                st.warning("⚠️ No related records found for this error.")
        
        # Edit summary and download section
        st.header("💾 Export Modified Data")
        
        if st.session_state.edit_history:
            # Show edit summary with delete buttons
            summary_df = create_edit_summary()
            
            # Download buttons
            col1, col2 = st.columns(2)
            
            
            with col1:
                # Download edit summary
                if st.session_state.edit_history:
                    # Create export dataframe directly from edit history
                    export_data = []
                    for edit in st.session_state.edit_history:
                        # Get descriptive column name
                        sheet_name = edit['sheet']
                        tech_name = edit['column']
                        descriptive_name = tech_name
                        
                        if sheet_name in st.session_state.header_mappings:
                            mapping = st.session_state.header_mappings[sheet_name]
                            if tech_name in mapping:
                                descriptive_name = mapping[tech_name]
                        
                        export_data.append({
                            'Sheet': sheet_name,
                            'Column': descriptive_name,
                            'Technical Name': tech_name,
                            'Row': edit['row'],
                            'Old Value': edit['old_value'],
                            'New Value': edit['new_value'],
                            'Error Reference': edit.get('error_reference', ''),
                            'Timestamp': edit['timestamp']
                        })
                    
                    export_df = pd.DataFrame(export_data)
                    csv_summary = export_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Edit Summary (CSV)",
                        data=csv_summary,
                        file_name="edit_summary.csv",
                        mime="text/csv"
                    )
            
            with col2:
                # Download modified XML
                if st.session_state.original_xml_content:
                    # Only update XML when the download button is clicked
                    if st.button("Prepare XML for Download"):
                        modified_xml, changes = update_xml_with_changes(
                            st.session_state.original_xml_content,
                            st.session_state.modified_dataframes
                        )
                        
                        if modified_xml:
                            st.session_state.download_xml = modified_xml
                            st.success(f"✅ {len(changes)} changes applied to XML")
                        else:
                            st.warning("No changes to apply to XML")
                    
                    # Only show the download button if we have prepared XML
                    if 'download_xml' in st.session_state:
                        st.download_button(
                            label="📥 Download Modified XML",
                            data=st.session_state.download_xml,
                            file_name="modified_migration_data.xml",
                            mime="application/xml"
                        )
        else:
            st.info("No edits made yet. Edit records above to enable download options.")
        
        # Worksheet relationships section
        with st.expander("📊 Data Structure Relationships"):
            st.markdown("""
            ### How the Worksheets Are Connected:
            
            - **Header Data (S_EKKO)**: Main purchase order record
            - **Item Data (S_EKPO)**: Line items linked to header via `EBELN`
            - **Account Assignment (S_EKKN)**: Financial data linked to items via `EBELN` and `EBELP`
            - **Schedule Line (S_EKET)**: Delivery schedules linked to items via `EBELN` and `EBELP`
            - **Header Texts (S_EKKO_TEXT)**: Text information for the PO header
            - **Item Texts (S_EKPO_TEXT)**: Text information for line items
            
            ### Common Error Patterns:
            - **Missing Purchasing Group**: Check `EKGRP` field in Header Data
            - **Material Text Issues**: Check `TXZ01` field in Item Data
            - **Account Assignment Issues**: Check `KNTTP` field in Account Assignment
            - **Plant Issues**: Check `WERKS` field in Item Data
            - **Delivery Date Issues**: Check `LDATE` in Item Data or `EINDT` in Schedule Line
            """)

if __name__ == "__main__":
    main()
